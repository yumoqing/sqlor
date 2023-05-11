from appPublic.jsonConfig import getConfig
import openpyxl as xlsx
import asyncio
from sqlor.dbpools import DBPools
import aiomysql
import aiopg
from typeconv import convrec

class CBObject:
	def __init__(self,db, name):
		self.db = db
		self.tbl = name

	async def handle(self,ws):
		db = DBPools()
		async with db.sqlorContext(self.db) as sor:
			delete_sql = "TRUNCATE TABLE %s" % self.tbl
			await sor.sqlExe(delete_sql, {})
			info = await sor.I(self.tbl)
			for rec in getRecord(ws):
				r = [ v for v in rec.values() if v is not None ]
				if len(r) == 0:
					continue
				rec = convrec(info, rec)
				await sor.C(self.tbl, rec)

def getRecord(ws):
	names = []
	types = []
	for i,r in enumerate(ws.rows):
		if i==0:
			for j,c in enumerate(r):
				nt = c.value.split(':')
				if len(nt) < 2:
					nt.append('')
				n,t = nt
				names.append(n)
				types.append(t)
		else:
			dic = {}
			for j,c in enumerate(r):
				v = c.value
				dic[names[j]] = v
			yield dic

async def loaddatainexcel(xlsxfile):
	wb = xlsx.load_workbook(xlsxfile)
	print(f'{wb.sheetnames=}')
	dbname = [ i[2:-2] for i in wb.sheetnames if i.startswith('__')][0]
	print(f'{dbname=}')
	for name in wb.sheetnames:
		if name.startswith('__'):
			continue
		ws = wb[name]
		cbobj = CBObject(dbname, name)
		await cbobj.handle(ws)

if __name__ == '__main__':
	import sys
	import os
	p = os.getcwd()
	config = getConfig(p)
	print(f'{config.databases=},cwd={p}')
	DBPools(config.databases)
	if len(sys.argv) < 2:
		print('%s xlsxfile' % sys.argv[0])
		sys.exit(1)
	loop = asyncio.get_event_loop()
	loop.run_until_complete(loaddatainexcel(sys.argv[1]))
